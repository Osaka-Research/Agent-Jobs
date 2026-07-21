"""
admin.py — sqlite log + xlsx export + telegram bot delivery.

Every successful scrape is logged here and (if TELEGRAM_BOT_TOKEN +
TELEGRAM_CHAT_ID are set) the xlsx is sent to that chat. one chat, no
subscribers, no broadcast, no polling.

Endpoints:
  POST /api/admin/log-search   — internal: called by main.py after each scrape
  GET  /api/admin/export      — admin: pull full xlsx of all logged searches
  GET  /api/admin/stats        — admin: row counts and last-search metadata

Env:
  TELEGRAM_BOT_TOKEN  — bot token for sending xlsx
  TELEGRAM_CHAT_ID    — chat id to send to
  SQLITE_PATH         — db file path (default /data/searches.db)
"""
from __future__ import annotations

from typing import Any

import io
import logging
import os
import sqlite3
import threading
from datetime import datetime, timezone
from pathlib import Path

import httpx
import openpyxl
from fastapi import APIRouter, HTTPException, Response
from pydantic import BaseModel, Field

log = logging.getLogger("agent-jobs.admin")

router = APIRouter(prefix="/api/admin", tags=["admin"])

DB_PATH = Path(os.getenv("SQLITE_PATH", "/data/searches.db"))
DB_PATH.parent.mkdir(parents=True, exist_ok=True)
TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "")
TELEGRAM_CHAT_ID = os.getenv("TELEGRAM_CHAT_ID", "")

_db_lock = threading.Lock()


def _connect() -> sqlite3.Connection:
    conn = sqlite3.connect(str(DB_PATH), check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn


def _init_db() -> None:
    with _db_lock:
        conn = _connect()
        try:
            conn.executescript(
                """
                CREATE TABLE IF NOT EXISTS searches (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    created_at TEXT NOT NULL,
                    search_term TEXT NOT NULL,
                    location TEXT NOT NULL,
                    sites TEXT NOT NULL,
                    hours_old INTEGER NOT NULL,
                    results_wanted INTEGER NOT NULL,
                    job_count INTEGER NOT NULL,
                    ok INTEGER NOT NULL,
                    duration_seconds REAL,
                    geo_lat REAL,
                    geo_lng REAL,
                    geo_accuracy REAL
                );
                CREATE INDEX IF NOT EXISTS idx_searches_created_at
                    ON searches(created_at DESC);

                CREATE TABLE IF NOT EXISTS jobs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    search_id INTEGER NOT NULL REFERENCES searches(id) ON DELETE CASCADE,
                    title TEXT,
                    company TEXT,
                    location TEXT,
                    site TEXT,
                    url TEXT,
                    date_posted TEXT,
                    salary_min INTEGER,
                    salary_max INTEGER,
                    salary_currency TEXT,
                    interval TEXT,
                    is_remote INTEGER,
                    job_type TEXT
                );
                CREATE INDEX IF NOT EXISTS idx_jobs_search_id ON jobs(search_id);

                CREATE TABLE IF NOT EXISTS subscribers (
                    chat_id INTEGER PRIMARY KEY,
                    first_name TEXT,
                    username TEXT,
                    subscribed_at TEXT NOT NULL,
                    last_seen_at TEXT NOT NULL
                );

                CREATE TABLE IF NOT EXISTS telegram_sessions (
                    search_id INTEGER PRIMARY KEY,
                    chat_id INTEGER NOT NULL,
                    message_id INTEGER NOT NULL,
                    opened_count INTEGER NOT NULL DEFAULT 0,
                    resume_count INTEGER NOT NULL DEFAULT 0,
                    apply_count INTEGER NOT NULL DEFAULT 0,
                    first_opened_title TEXT,
                    first_opened_company TEXT,
                    first_applied_title TEXT,
                    first_applied_company TEXT,
                    first_resumed_title TEXT,
                    first_resumed_company TEXT,
                    last_event_at TEXT,
                    created_at TEXT NOT NULL
                );
                """
            )
            conn.commit()
        finally:
            conn.close()


def add_subscriber(chat_id: int, first_name: str | None, username: str | None) -> None:
    """upsert a subscriber row when they /start the bot."""
    now = datetime.now(timezone.utc).isoformat()
    with _db_lock:
        conn = _connect()
        try:
            conn.execute(
                """INSERT INTO subscribers (chat_id, first_name, username, subscribed_at, last_seen_at)
                   VALUES (?, ?, ?, ?, ?)
                   ON CONFLICT(chat_id) DO UPDATE SET
                       first_name = excluded.first_name,
                       username = excluded.username,
                       last_seen_at = excluded.last_seen_at""",
                (chat_id, first_name, username, now, now),
            )
            conn.commit()
        finally:
            conn.close()


def list_subscribers() -> list[int]:
    conn = _connect()
    try:
        return [r["chat_id"] for r in conn.execute("SELECT chat_id FROM subscribers").fetchall()]
    finally:
        conn.close()


_init_db()


class LogJob(BaseModel):
    title: str | None = None
    company: str | None = None
    location: str | None = None
    site: str | None = None
    url: str | None = None
    date_posted: str | None = None
    salary_min: float | None = None
    salary_max: float | None = None
    salary_currency: str | None = None
    interval: str | None = None
    is_remote: bool | None = None
    job_type: str | None = None


class LogSearch(BaseModel):
    search_term: str
    location: str = ""
    sites: list[str] = Field(default_factory=list)
    hours_old: int = 168
    results_wanted: int = 50
    job_count: int = 0
    ok: bool = True
    duration_seconds: float | None = None
    geo: dict | None = Field(None, description="optional gps fix {lat, lng, accuracy}")
    jobs: list[LogJob] = Field(default_factory=list)


class LogEvent(BaseModel):
    """user-action event from the dashboard. backend increments counters
    and edits the corresponding telegram session message in place."""
    search_id: int
    event: str = Field(..., description="one of: open, resume, apply")
    job_id: int | None = None
    job_title: str | None = None
    job_company: str | None = None
    job_url: str | None = None


def _build_searches_workbook() -> openpyxl.Workbook:
    conn = _connect()
    try:
        searches = conn.execute("SELECT * FROM searches ORDER BY id DESC").fetchall()
        jobs = conn.execute(
            """SELECT j.id, j.search_id, j.title, j.company, j.location, j.site,
                       j.url, j.date_posted, j.salary_min, j.salary_max,
                       j.salary_currency, j.interval, j.is_remote, j.job_type,
                       s.created_at AS search_created_at,
                       s.search_term, s.location AS search_location
                 FROM jobs j JOIN searches s ON j.search_id = s.id
                ORDER BY j.id DESC"""
        ).fetchall()
    finally:
        conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "searches"
    headers = ["id", "created_at", "search_term", "location", "sites",
               "hours_old", "results_wanted", "job_count", "ok", "duration_seconds"]
    ws.append(headers)
    _style_header(ws)
    for r in searches:
        ws.append([r[h] for h in headers])

    ws2 = wb.create_sheet("jobs")
    headers2 = ["id", "search_id", "search_created_at", "search_term", "search_location",
                "title", "company", "location", "site", "url", "date_posted",
                "salary_min", "salary_max", "salary_currency", "interval", "is_remote", "job_type"]
    ws2.append(headers2)
    _style_header(ws2)
    for j in jobs:
        ws2.append([j[h] for h in headers2])

    for sheet in (ws, ws2):
        for col_idx, _ in enumerate(sheet[1], start=1):
            col_letter = openpyxl.utils.cell.get_column_letter(col_idx)
            sheet.column_dimensions[col_letter].width = 28
    return wb


def _build_single_search_xlsx(payload: LogSearch) -> bytes:
    """build a tiny xlsx with just the search metadata + its jobs."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "search"
    ws.append(["search_term", payload.search_term])
    ws.append(["location", payload.location])
    ws.append(["sites", ",".join(payload.sites)])
    ws.append(["hours_old", payload.hours_old])
    ws.append(["results_wanted", payload.results_wanted])
    ws.append(["job_count", payload.job_count])
    ws.append(["ok", payload.ok])
    if payload.duration_seconds is not None:
        ws.append(["duration_seconds", payload.duration_seconds])
    ws.append([])
    hdr = ["title", "company", "location", "site", "url", "salary", "remote", "type"]
    ws.append(hdr)
    for cell in ws[ws.max_row]:
        cell.font = openpyxl.styles.Font(bold=True)
    for j in payload.jobs:
        sal = ""
        if j.salary_min and j.salary_max:
            sal = f"{j.salary_currency or ''} {j.salary_min}-{j.salary_max} {j.interval or ''}".strip()
        elif j.salary_min:
            sal = f"{j.salary_currency or ''} {j.salary_min} {j.interval or ''}".strip()
        ws.append([j.title, j.company, j.location, j.site, j.url, sal,
                   "yes" if j.is_remote else "", j.job_type])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_caption(payload: LogSearch, created_at: str) -> str:
    """structured caption — each line is parseable by a downstream cloud caller.

    Format:
      🔍 <role> @ <location>
      📍 <lat>,<lng> (±<accuracy_m>m)        # only if payload.geo is set
      📊 <count> jobs · <sites> · within <hours>h
      ⏱ <duration>s · <created_at_utc>
    """
    lines = []
    role = payload.search_term
    loc = f" @ {payload.location}" if payload.location else ""
    lines.append(f"🔍 {role!r}{loc}")
    if payload.geo:
        lat = payload.geo.get("lat")
        lng = payload.geo.get("lng")
        acc = payload.geo.get("accuracy")
        if lat is not None and lng is not None:
            acc_str = f" (±{acc}m)" if acc is not None else ""
            lines.append(f"📍 {lat:.4f},{lng:.4f}{acc_str}")
    sites = ",".join(payload.sites) if payload.sites else ""
    lines.append(f"📊 {payload.job_count} jobs · {sites} · within {payload.hours_old}h")
    dur = f"{payload.duration_seconds}s" if payload.duration_seconds is not None else "?"
    lines.append(f"⏱ {dur} · {created_at}")
    return "\n".join(lines)


def _build_session_text(payload: LogSearch, created_at: str,
                        opened: int = 0, resume: int = 0, apply: int = 0,
                        first_opened: tuple[str, str] | None = None,
                        first_applied: tuple[str, str] | None = None,
                        first_resumed: tuple[str, str] | None = None) -> str:
    """text-only body of the session message that we keep editing in place
    as the user clicks cards. starts with the search summary and a
    placeholder action-counter line that gets updated.

    first_opened/applied/resumed are (title, company) tuples — the first
    job the user took that action on. only set when that action has
    happened at least once, so the line is meaningful.
    """
    lines = [_build_caption(payload, created_at)]
    lines.append("─" * 24)
    lines.append(
        f"👆 {opened} opened  ·  📄 {resume} resumes  ·  🔗 {apply} applied"
    )
    # show the first job per action — high-signal info for a cloud caller
    if first_opened:
        title, company = first_opened
        lines.append(f"first opened:   \"{title}\" @ {company}")
    if first_applied:
        title, company = first_applied
        lines.append(f"first applied:  \"{title}\" @ {company}")
    if first_resumed:
        title, company = first_resumed
        lines.append(f"first resumed:  \"{title}\" @ {company}")
    return "\n".join(lines)


def _style_header(ws) -> None:
    fill = openpyxl.styles.PatternFill(start_color="3ddc97", end_color="3ddc97", fill_type="solid")
    font = openpyxl.styles.Font(bold=True, color="0b0f0d")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font


async def _send_xlsx_to_bot(xlsx_bytes: bytes, filename: str, caption: str) -> bool:
    if not TELEGRAM_BOT_TOKEN or not TELEGRAM_CHAT_ID:
        log.info("telegram delivery skipped (TELEGRAM_BOT_TOKEN or TELEGRAM_CHAT_ID not set)")
        return False
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument"
    try:
        async with httpx.AsyncClient(timeout=30) as client:
            files = {"document": (filename, xlsx_bytes,
                                  "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
            data = {"chat_id": TELEGRAM_CHAT_ID, "caption": caption[:1024]}
            r = await client.post(url, files=files, data=data)
            if r.status_code != 200:
                log.warning(f"telegram sendDocument failed: {r.status_code} {r.text[:200]}")
                return False
            return True
    except Exception as e:
        log.exception(f"telegram sendDocument exception: {e}")
        return False


async def _send_text_to_bot(text: str) -> int | None:
    """send a plain text message and return its message_id, or None on failure."""
    if not TELEGRAM_BOT_TOKEN or not TELEGRAM_CHAT_ID:
        return None
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
    try:
        async with httpx.AsyncClient(timeout=15) as client:
            r = await client.post(url, json={
                "chat_id": TELEGRAM_CHAT_ID,
                "text": text[:4096],
            })
            if r.status_code != 200:
                log.warning(f"sendMessage failed: {r.status_code} {r.text[:200]}")
                return None
            data = r.json()
            return data.get("result", {}).get("message_id")
    except Exception as e:
        log.exception(f"sendMessage exception: {e}")
        return None


async def _edit_message(message_id: int, text: str) -> bool:
    """edit a previously-sent message in place. returns True on success."""
    if not TELEGRAM_BOT_TOKEN or not TELEGRAM_CHAT_ID or not message_id:
        return False
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/editMessageText"
    try:
        async with httpx.AsyncClient(timeout=15) as client:
            r = await client.post(url, json={
                "chat_id": TELEGRAM_CHAT_ID,
                "message_id": message_id,
                "text": text[:4096],
            })
            if r.status_code != 200:
                log.debug(f"editMessageText failed: {r.status_code} {r.text[:120]}")
                return False
            return True
    except Exception as e:
        log.debug(f"editMessageText exception: {e}")
        return False


@router.post("/log-search")
async def log_search(payload: LogSearch) -> dict:
    """internal endpoint: log search + jobs + send xlsx to bot."""
    created_at = datetime.now(timezone.utc).isoformat()
    sites_csv = ",".join(payload.sites)

    with _db_lock:
        conn = _connect()
        try:
            geo_lat = payload.geo.get("lat") if payload.geo else None
            geo_lng = payload.geo.get("lng") if payload.geo else None
            geo_acc = payload.geo.get("accuracy") if payload.geo else None
            cur = conn.execute(
                """INSERT INTO searches
                       (created_at, search_term, location, sites, hours_old,
                        results_wanted, job_count, ok, duration_seconds,
                        geo_lat, geo_lng, geo_accuracy)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (created_at, payload.search_term, payload.location, sites_csv,
                 payload.hours_old, payload.results_wanted, payload.job_count,
                 1 if payload.ok else 0, payload.duration_seconds,
                 geo_lat, geo_lng, geo_acc),
            )
            search_id = cur.lastrowid
            for j in payload.jobs:
                conn.execute(
                    """INSERT INTO jobs
                           (search_id, title, company, location, site, url,
                            date_posted, salary_min, salary_max, salary_currency,
                            interval, is_remote, job_type)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (search_id, j.title, j.company, j.location, j.site, j.url,
                     j.date_posted, j.salary_min, j.salary_max, j.salary_currency,
                     j.interval, 1 if j.is_remote else 0, j.job_type),
                )
            conn.commit()
        finally:
            conn.close()

    delivered = False
    session_msg_id = None
    if payload.ok and payload.jobs:
        xlsx = _build_single_search_xlsx(payload)
        fname = f"search-{search_id}-{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}.xlsx"
        caption = _build_caption(payload, created_at)
        delivered = await _send_xlsx_to_bot(xlsx, fname, caption)
        # also send the session-tracker text message that we'll keep editing
        # as the user clicks cards. store message_id for later edits.
        session_text = _build_session_text(payload, created_at, 0, 0, 0)
        session_msg_id = await _send_text_to_bot(session_text)

    # persist the session row regardless of telegram success — if telegram
    # is down, the search is still recoverable from sqlite + we can edit later.
    if session_msg_id is not None:
        try:
            with _db_lock:
                conn = _connect()
                try:
                    conn.execute(
                        """INSERT INTO telegram_sessions
                               (search_id, chat_id, message_id, created_at)
                           VALUES (?, ?, ?, ?)""",
                        (search_id, TELEGRAM_CHAT_ID, session_msg_id, created_at),
                    )
                    conn.commit()
                finally:
                    conn.close()
        except Exception:
            log.exception("failed to persist telegram_sessions row (non-fatal)")

    return {"ok": True, "search_id": search_id, "logged_jobs": len(payload.jobs),
            "telegram_delivered": delivered,
            "telegram_message_id": session_msg_id}


@router.get("/stats")
async def stats() -> dict:
    conn = _connect()
    try:
        n_searches = conn.execute("SELECT COUNT(*) AS n FROM searches").fetchone()["n"]
        n_jobs = conn.execute("SELECT COUNT(*) AS n FROM jobs").fetchone()["n"]
        last = conn.execute(
            "SELECT created_at, search_term, location, job_count FROM searches "
            "ORDER BY id DESC LIMIT 1"
        ).fetchone()
        return {
            "searches_count": n_searches,
            "jobs_count": n_jobs,
            "last_search": dict(last) if last else None,
            "telegram_enabled": bool(TELEGRAM_BOT_TOKEN and TELEGRAM_CHAT_ID),
            "subscribers_count": len(list_subscribers()),
            "bot_token_set": bool(TELEGRAM_BOT_TOKEN),
            "chat_id_set": bool(TELEGRAM_CHAT_ID),
        }
    finally:
        conn.close()


@router.post("/event")
async def log_event(payload: LogEvent) -> dict:
    """dashboard POSTs here when the user clicks open/resume/apply on a card.
    we increment counters in telegram_sessions, set the first-X columns
    on first occurrence, and edit the corresponding message in place.
    the persistent message keeps the same message_id — only its text
    updates. no per-event chat spam."""
    if payload.event not in ("open", "resume", "apply"):
        raise HTTPException(status_code=400, detail=f"unknown event '{payload.event}'")

    now_iso = datetime.now(timezone.utc).isoformat()
    with _db_lock:
        conn = _connect()
        try:
            row = conn.execute(
                "SELECT chat_id, message_id, opened_count, resume_count, apply_count, "
                "first_opened_title, first_opened_company, "
                "first_applied_title, first_applied_company, "
                "first_resumed_title, first_resumed_company "
                "FROM telegram_sessions WHERE search_id = ?",
                (payload.search_id,),
            ).fetchone()
            if row is None:
                return {"ok": True, "edited": False, "reason": "no session message"}

            # determine which column to update for the first-action record
            first_title_col = {"open": "first_opened_title", "apply": "first_applied_title", "resume": "first_resumed_title"}[payload.event]
            first_company_col = {"open": "first_opened_company", "apply": "first_applied_company", "resume": "first_resumed_company"}[payload.event]
            count_col = {"open": "opened_count", "apply": "apply_count", "resume": "resume_count"}[payload.event]

            # bump counter always; set first-X columns only if currently null
            set_clauses = [f"{count_col} = {count_col} + 1", "last_event_at = ?"]
            params: list[Any] = [now_iso]

            if payload.job_title and row[first_title_col] is None:
                set_clauses.append(f"{first_title_col} = ?")
                params.append(payload.job_title)
                if payload.job_company is not None:
                    set_clauses.append(f"{first_company_col} = ?")
                    params.append(payload.job_company)

            params.append(payload.search_id)
            conn.execute(
                f"UPDATE telegram_sessions SET {', '.join(set_clauses)} WHERE search_id = ?",
                params,
            )
            conn.commit()
            # re-read
            row = conn.execute(
                "SELECT chat_id, message_id, opened_count, resume_count, apply_count, "
                "first_opened_title, first_opened_company, "
                "first_applied_title, first_applied_company, "
                "first_resumed_title, first_resumed_company "
                "FROM telegram_sessions WHERE search_id = ?",
                (payload.search_id,),
            ).fetchone()
        finally:
            conn.close()

    # build the updated message text from the original search + new state
    with _db_lock:
        conn = _connect()
        try:
            search_row = conn.execute(
                "SELECT * FROM searches WHERE id = ?", (payload.search_id,)
            ).fetchone()
        finally:
            conn.close()
    if search_row is None:
        return {"ok": True, "edited": False, "reason": "search not found"}

    payload_for_text = LogSearch(
        search_term=search_row["search_term"],
        location=search_row["location"],
        sites=search_row["sites"].split(",") if search_row["sites"] else [],
        hours_old=search_row["hours_old"],
        results_wanted=search_row["results_wanted"],
        job_count=search_row["job_count"],
        ok=bool(search_row["ok"]),
        duration_seconds=search_row["duration_seconds"],
        geo=(
            {"lat": search_row["geo_lat"], "lng": search_row["geo_lng"],
             "accuracy": search_row["geo_accuracy"]}
            if search_row["geo_lat"] is not None else None
        ),
    )
    text = _build_session_text(
        payload_for_text,
        search_row["created_at"],
        opened=row["opened_count"],
        resume=row["resume_count"],
        apply=row["apply_count"],
        first_opened=(row["first_opened_title"], row["first_opened_company"]) if row["first_opened_title"] else None,
        first_applied=(row["first_applied_title"], row["first_applied_company"]) if row["first_applied_title"] else None,
        first_resumed=(row["first_resumed_title"], row["first_resumed_company"]) if row["first_resumed_title"] else None,
    )
    edited = await _edit_message(row["message_id"], text)

    return {"ok": True, "edited": edited,
            "counters": {"opened": row["opened_count"],
                         "resume": row["resume_count"],
                         "apply": row["apply_count"]}}


@router.get("/subscribers")
async def subscribers() -> dict:
    """return list of registered subscribers (chats that /start'd the bot)."""
    conn = _connect()
    try:
        rows = conn.execute(
            "SELECT chat_id, first_name, username, subscribed_at, last_seen_at "
            "FROM subscribers ORDER BY subscribed_at DESC"
        ).fetchall()
        return {
            "count": len(rows),
            "subscribers": [dict(r) for r in rows],
        }
    finally:
        conn.close()


@router.get("/export")
async def export() -> Response:
    wb = _build_searches_workbook()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"agent-jobs-{datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')}.xlsx"
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
